import openpyxl
import os

def calcAvgPWD(sheet, num):
    row1 = 3
    addFactor = 1.0 / num
    runningSum = 0
    numComps = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        for row2 in sheet.iter_rows(min_row = row1, values_only=True):
            sum = 0.0
            numComps+=1
            for i in range(num):
                #If different
                if row[9+i] != row2[9+i]:
                    sum += addFactor
            runningSum += sum
        row1+=1
    if numComps != 0:
        return float(runningSum / numComps)
    return 0 

##Get metadata from key row
#Return: tuple with (num_reads, unique_reads, 
#                   avgPWD, fcpg, and an array of avg X Vals)
def calcResults(sheet, num, important):
    num_reads = sheet.max_row - 1
    if num_reads == 0:
        return (0, 0, 0, 0, [0 for i in range(num)])
    distinct_elems = set()
    important_sum = 0
    Xvals = [0 for i in  range(num)]
    #Main loop
    for row in sheet.iter_rows(min_row=2, values_only=True):
        #Check for unique vals
        if row[8] not in distinct_elems:
            distinct_elems.add(row[8])
        important_sum += int(row[8+important])
        for i in range(num):
            Xvals[i] += int(row[9+i])
    #Average out Xvals
    for i in range(num):
        Xvals[i] = float(Xvals[i] / num_reads)
    avgPWD = calcAvgPWD(sheet, num)
    return (num_reads, len(distinct_elems), avgPWD, float(important_sum / num_reads), Xvals)

def generateOutput(name):
    inp = openpyxl.load_workbook(filename='input/{}.xlsx'.format(name))
    key = openpyxl.load_workbook(filename="keys/{}-key.xlsx".format(name))
    outp = openpyxl.Workbook()
    outfilename = "output/{}-output.xlsx".format(name)

    ##Initial output creation
    outsheet = outp.active
    keysheet = key.active
    maxsites = 0

    ##Copy key to output
    for row in keysheet:
        for cell in row:
            outsheet[cell.coordinate].value = cell.value

    for cell in keysheet['C'][2:]:
        if int(cell.value) > maxsites:
            maxsites = int(cell.value)

    #Add other output header cols
    outsheet['G1'].value = 'number of reads'
    outsheet['H1'].value = 'number of unique reads'
    outsheet['I1'].value = 'average PWD'
    outsheet['J1'].value = 'fCpG'

    startcol = 11
    for i in range(1, maxsites+1):
        outsheet.cell(row = 1, column=startcol, value="X{}".format(i))
        startcol+=1

    for rowIdx in range(2, outsheet.max_row+1):
        #Sheet 0 of data will be row 2 of input 
        #Accounting for 0 indexing and header rows
        numsites = int(outsheet['C{}'.format(rowIdx)].value)
        important = int(outsheet['D{}'.format(rowIdx)].value)
        result = calcResults(inp.worksheets[rowIdx-2], numsites, important)
        #Write results to output
        outsheet['G{}'.format(rowIdx)] = result[0]
        outsheet['H{}'.format(rowIdx)] = result[1]
        outsheet['I{}'.format(rowIdx)] = result[2]
        outsheet['J{}'.format(rowIdx)] = result[3]
        for i in range(numsites):
            outsheet.cell(row = rowIdx, column=11+i, value=result[4][i])
    
    #Part 2: Add raw data
    outp.create_sheet("Raw")
    outsheet = outp['Raw']

    #Label columns for raw sheet
    outsheet['A1'] = 'Calum70'
    outsheet['B1'] = 'name'
    outsheet['C1'] = 'CpG in Amplicon'
    outsheet['D1'] = 'epicsite'
    outsheet['E1'] = 'comments'
    outsheet['F1'] = 'direction'
    outsheet['G1'] = 'numnc'
    outsheet['H1'] = 'ncph'
    outsheet['I1'] = 'sample'
    outsheet['J1'] = 'amplicon'
    outsheet['K1'] = 'chr'
    outsheet['L1'] = 'start'
    outsheet['M1'] = 'readseq'
    outsheet['N1'] = 'naCpG'
    outsheet['O1'] = 'htype'
    startcol = 16
    for i in range(1, maxsites+1):
        outsheet.cell(row = 1, column=startcol, value="X{}".format(i))
        startcol+=1
    index=1
    for ws in inp.worksheets:
        index+=1
        keyrow = keysheet.iter_rows(min_row = index, max_row = index, min_col=1, max_col=6)
        for row in keyrow:
            keyRowStruct = [cell.value for cell in row]
        for row in ws.iter_rows(min_row=2):
            rowStruct = [cell.value for cell in row]
            finalRow = [*keyRowStruct, *rowStruct]
            outsheet.append(finalRow)    

    outp.save(filename=outfilename)

namelist = open("list.txt", "r")
for name in namelist:
    generateOutput(name)